# -*- coding:utf-8 -*-
"""
@function：读取Excel中的测试数据
"""
import openpyxl
from common import logger

logger = logger.get_logger('do_excel')
class Case:
    # 测试用例封装类，每一个实例表示一条测试用例
    def __init__(self):
        self.id = None # 测试用例id
        self.url = None # 测试用例url
        self.data = None # 测试用例data
        self.title = None # 测试用例title
        self.method = None # 测试用例method
        self.expected = None # 测试用例expected
        self.actual = None # 测试用例actual
        self.result = None # 测试用例result

class DoExcel:
    '''读取表格中的测试数据'''
    def __init__(self,file_name):
        try:
            self.file_name = file_name # 要读取的文件名
            self.workbook = openpyxl.load_workbook(filename=file_name) # 实例化一个workbook
        except FileNotFoundError as e:
            # 文件未找到进行异常处理
            logger.error('{0} not found, please check file path.'.format(file_name))
            raise e # 抛出异常
    def get_cases(self,sheet_name): #从表格中读取测试用例的方法
        sheet = self.workbook[sheet_name] # 获取sheet
        max_row = sheet.max_row # 获取sheet最大行
        cases = [] # 定义一个列表，用来存放读取的测试用例
        for r in range(2,max_row+1):
            case = Case() # 实例化一个case对象，用来存放测试用例
            case.id = sheet.cell(row = r,column = 1).value  # 取第r行，第1列的case_id值
            case.title = sheet.cell(row = r,column = 2).value  # 取第r行，第2列的title值
            case.url = sheet.cell(row=r, column=3).value  # 取第r行，第3列的url值
            case.data = sheet.cell(row=r, column=4).value  # 取第r行，第4列的data值
            case.method = sheet.cell(row=r, column=5).value  # 取第r行，第5列的method值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6列的expected值
            if type(case.expected) == int: # excepted中只有code代码的，转化成str，因为request返回中的code代码是str类型
                case.expected = str(case.expected)
            cases.append(case) # 将case添加到列表cases中
        return cases # for循环结束后返回cases列表
    def write_result(self,sheet_name,row,actual,result): #将测试用例的结果写回到表格中
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row ,column=7).value = actual # 写入实际结果
        sheet.cell(row=row, column=8).value = result  # 写入执行结果
        self.workbook.save(filename=self.file_name) # 保存

if __name__ == '__main__':
    from common import constants
    do_excel = DoExcel(constants.case_file)
    cases = do_excel.get_cases('register')
    for case in cases:
        print(case.data)
        print(type(case.data))
        print(case.id,type(case.id))




